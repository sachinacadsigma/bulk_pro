# Q&A Chatbot
# from langchain.llms import OpenAI

from dotenv import load_dotenv

load_dotenv()  # take environment variables from .env.

import streamlit as st
import os
import time
from PIL import Image
import json
import pandas as pd
from propelauth import auth
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
import anthropic
from anthropic.types.message_create_params import MessageCreateParamsNonStreaming
from anthropic.types.messages.batch_create_params import Request
import base64

# st.set_page_config(page_title="Invoice Extractor")

user = auth.get_user()

if user is None:
    st.error("Unauthorized")
    st.stop()

if "user_email" not in st.session_state:
    file_dir = os.path.join(os.getcwd(), "excel_files")
    st.session_state["user_email"] = user.email
    user_dir_name = user.email.split("@")[0]
    user_dir_path = os.path.join(file_dir, user_dir_name)
    if user_dir_name not in os.listdir(file_dir):
        os.mkdir(user_dir_path)
    st.session_state["user_dir_path"] = user_dir_path
    st.session_state["user_dir_name"] = user_dir_name

system_prompt = None
with open("System Prompt.txt", "r") as f:
    system_prompt = f.read()

client = anthropic.Client(api_key=os.getenv("ANTHROPIC_API_KEY"))


def create_batch(files):
    reqs = []
    for idx, file in enumerate(files, 1):
        file_data = base64.standard_b64encode(file.getvalue()).decode("utf-8")
        file_type = "document" if file.type == "application/pdf" else "image"
        req = Request(
            custom_id=f"inv-req-{idx}",
            params=MessageCreateParamsNonStreaming(
                model="claude-3-5-sonnet-latest",
                system=system_prompt,
                max_tokens=4096,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": file_type,
                                "source": {
                                    "type": "base64",
                                    "media_type": file.type,
                                    "data": file_data,
                                },
                            },
                            {"type": "text", "text": "This is the invoice."},
                        ],
                    }
                ],
            ),
        )
        reqs.append(req)

    message_batch = client.messages.batches.create(requests=reqs)
    return message_batch


def convert_batch_results_to_df(message_batch, header_df, items_df):
    item_id = 1
    for header_id, result in enumerate(
        client.messages.batches.results(message_batch_id=message_batch.id), 1
    ):
        match result.result.type:
            case "succeeded":
                try:
                    json_data = json.loads(result.result.message.content[0].text)

                    header_data = json_data["HEADER"]
                    header_data["ID"] = header_id
                    header_df.loc[len(header_df)] = header_data

                    items_data = json_data["ITEMS"]
                    for item_data in items_data:
                        item_data["ID"] = item_id
                        item_data["HEADER_ID"] = header_id
                        items_df.loc[len(items_df)] = item_data
                        item_id += 1

                except Exception as e:
                    st.error(
                        f"Some error occured in file {header_id}... Sorry for the inconvinience..."
                    )
            case "errored":
                st.error(
                    f"Some error occured in file {header_id}... Sorry for the inconvinience..."
                )
            case "expired":
                print("Request expired...")


def create_excel_file(header_df, items_df, book_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    table_caption_font = Font(name="Calibri", size=14, bold=True)
    ws["A1"] = "HEADER TABLE"
    ws["A1"].font = table_caption_font
    # Write the header_df DataFrame to the worksheet
    for row in dataframe_to_rows(header_df, index=False, header=True):
        ws.append(row)
        # Define the range for the table
    table_range = f"A2:R{len(header_df) + 2}"
    # Create a table
    table = Table(displayName="HeaderTable", ref=table_range)
    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium13",  # You can change the style
        showRowStripes=True,
    )
    table.tableStyleInfo = style
    # Add the table to the worksheet
    ws.add_table(table)

    row_num = len(header_df) + 4
    ws[f"A{row_num}"] = "ITEMS TABLE"
    ws[f"A{row_num}"].font = table_caption_font
    # Write the header_df DataFrame to the worksheet
    for row in dataframe_to_rows(items_df, index=False, header=True):
        ws.append(row)
        # Define the range for the table
    table_range = f"A{row_num + 1}:R{len(items_df) + row_num + 1}"
    # Create a table
    table = Table(displayName="ItemsTable", ref=table_range)
    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium13",  # You can change the style
        showRowStripes=True,
    )
    table.tableStyleInfo = style
    # Add the table to the worksheet
    ws.add_table(table)
    wb.save(book_path)


@st.fragment
def download_file(book_path, file_name):
    with open(book_path, "rb") as f:
        st.download_button(
            label="Download excel file",
            data=f,
            file_name=f"{file_name}.xlsx",
            mime="application/vnd.ms-excel",
        )


##initialize our streamlit app
st.header("Invoice Extractor Application")
uploaded_files = st.file_uploader(
    "Choose an image or pdf...",
    type=["jpg", "jpeg", "png", "pdf"],
    accept_multiple_files=True,
)
image = ""
submit = None
if len(uploaded_files) > 0:
    submit = st.button("Submit")


## If ask button is clicked
if submit:
    message_batch = create_batch(uploaded_files)
    container = st.empty()
    container.info("Processing...", icon=":material/progress_activity:")
    while True:
        message_batch = client.messages.batches.retrieve(message_batch.id)
        if message_batch.processing_status == "ended":
            break

        container.info("Still Processing...", icon=":material/progress_activity:")
        time.sleep(10)

    header_df = pd.DataFrame(
        columns=[
            "ID",
            "INVOICE_NO",
            "INVOICE_DATE",
            "TOTAL_BEFORE_TAXES",
            "GST_CHARGES",
            "INVOICE_DISCOUNT",
            "INVOICE_TOTAL",
            "CURRENCY",
            "VENDOR_NAME",
            "VENDOR_ADDRESS",
            "VENDOR_GSTIN",
            "VENDOR_CIN",
            "VENDOR_PAN",
            "CUST_NAME",
            "CUST_ADDRESS",
            "CUST_GSTIN",
            "CUST_PAN",
            "CUST_CIN",
        ]
    )
    items_df = pd.DataFrame(
        columns=[
            "ID",
            "HEADER_ID",
            "ITEM_NAME",
            "ITEM_DESCRIPTION",
            "ITEM_QUANTITY",
            "UNIT_PRICE",
            "SKU_CODE",
            "HSN",
            "TOTAL_BEFORE_TAXES",
            "DISCOUNT",
            "SGST_RATE",
            "CGST_RATE",
            "IGST_RATE",
            "SGST_AMOUNT",
            "CGST_AMOUNT",
            "IGST_AMOUNT",
            "GRAND_TOTAL",
            "CURRENCY",
        ]
    )
    convert_batch_results_to_df(message_batch, header_df, items_df)
    try:
        file_name = (
            st.session_state["user_dir_name"]
            + " "
            + str(datetime.now().strftime("%d-%b-%Y %I_%M_%S_%p"))
        )
        book_path = os.path.join(st.session_state["user_dir_path"], f"{file_name}.xlsx")
        create_excel_file(header_df, items_df, book_path)
        container.empty()
        container.success("Processing completed successfully!!!")
        download_file(book_path, file_name)

    except Exception as e:
        st.error("Some error occured, sorry for the inconvenience. Try again later!")
